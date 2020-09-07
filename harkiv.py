from selenium import webdriver
import csv
from bs4 import BeautifulSoup as bs
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font

# путь к драйверу chrome
chromedriver = '/home/dony/chromedriver'
options = webdriver.ChromeOptions()
options.add_argument('headless')  # для открытия headless-браузера
options.add_argument('--no-sandbox')  # 
browser = webdriver.Chrome(executable_path=chromedriver, options=options)
url = 'https://hrk.aero/onlajjn-tablo'


# Запускаем хром для получения данных
browser.get(url)
# Получение HTML-содержимого
requiredHtml = browser.page_source
browser.quit()

# Метод для парсинга
def get_table(url):
    soup = bs(requiredHtml, 'html5lib')

    # ищем таблицы на сегодня(прилет, вылет)
    arrival_table = soup.find('table', id="table-arrival")
    departure_table = soup.find('table', id="table-departing")
    arrivals = parse(arrival_table)
    departures = parse(departure_table)
    write_excel(arrivals, departures)
    


# Парсим данные на одну из направлений
def parse(table):
    headers = []
    data = []
    # готовим Заголовки
    headers.append('дата')
    for th in table.findAll('th'):
        headers.append( th.text.replace('\n', ' ').strip())
    for tr in table.findAll('tr', class_='flight-item flight-today'):
        table_data = []
        table_data.append(str(date.today()))
        for td in tr.findAll('td'):
            table_data.append(td.text.replace('\n', ' ').strip())
        data.append(table_data)

    return  [headers] + data


# Записываем данные в csv файл
def write_file(data):
    with open( date.today().strftime('%Y-%m-%d') +'.csv', 'w') as csvfile:
        writer = csv.writer(csvfile)
        for row in data:
            writer.writerow(row)


# Запись в Ехзель
def write_excel(arrival,departure):
    wb = Workbook()
    ws = wb.active
    ws.title = "Arrivals"
    ws2 = wb.create_sheet("Departures")

    ws.sheet_properties.tabColor = "00ff11"
    ws2.sheet_properties.tabColor = "ff0000"
    
    write_array(arrival, ws)
    write_array(departure, ws2)
    wb.save(datetime.now().strftime('%d-%m-%Y') +'HARKOV.xlsx')



def write_array(data, ws):
    r = 1  
    for row in data:
        c = 1
        for value in row:
            ws.cell(row=r, column=c).value = value
            c += 1
        r += 1
    font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = font
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.auto_filter.ref = ws.dimensions


get_table(url)